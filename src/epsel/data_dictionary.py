"""Genera diccionarios de datos en Excel a partir de un DataFrame.

Perfila cada columna (tipo, completitud, cardinalidad, ejemplos), añade la
descripción curada y una bandera de calidad, y lo escribe a un .xlsx con
formato (hoja Resumen + hoja Diccionario).
"""
from __future__ import annotations

from pathlib import Path

import geopandas as gpd
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Paleta
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_TITLE_FONT = Font(color="1F4E78", bold=True, size=14)
_BAD_FILL = PatternFill("solid", fgColor="F8CBAD")    # vacía / constante
_WARN_FILL = PatternFill("solid", fgColor="FFE699")   # mayoría nula
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_COLS = ["Columna", "Descripción", "Tipo de dato", "% No nulos",
         "Nº únicos", "Ejemplos", "Observaciones"]


def _quality_flag(non_null_pct: float, nunique: int, n_rows: int, dtype_kind: str) -> str:
    if non_null_pct == 0:
        return "VACÍA (100% nulos)"
    if nunique <= 1:
        return "Constante (sin variabilidad)"
    if non_null_pct < 50:
        return "Mayoría nula"
    # "Identificador" solo para texto/enteros, no para medidas continuas (float)
    if nunique == n_rows and n_rows > 0 and dtype_kind in ("O", "i", "u"):
        return "Identificador (todos únicos)"
    return ""


def profile_dataframe(df: pd.DataFrame, descriptions: dict[str, str]) -> pd.DataFrame:
    """Devuelve la tabla del diccionario para `df`."""
    n_rows = len(df)
    rows = []
    for col in df.columns:
        if col == "geometry":
            continue
        s = df[col]
        non_null_pct = round(s.notna().mean() * 100, 1)
        nunique = int(s.nunique(dropna=True))
        ejemplos = ", ".join(str(x)[:30] for x in s.dropna().unique()[:3]) or "—"
        rows.append({
            "Columna": col,
            "Descripción": descriptions.get(col, ""),
            "Tipo de dato": str(s.dtype),
            "% No nulos": non_null_pct,
            "Nº únicos": nunique,
            "Ejemplos": ejemplos,
            "Observaciones": _quality_flag(non_null_pct, nunique, n_rows, s.dtype.kind),
        })
    return pd.DataFrame(rows, columns=_COLS)


def write_dictionary(
    profile: pd.DataFrame, meta: dict, out_path: Path,
) -> Path:
    """Escribe el diccionario a Excel con dos hojas: Resumen y Diccionario."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        # --- Hoja Resumen ---
        meta_df = pd.DataFrame(list(meta.items()), columns=["Campo", "Valor"])
        meta_df.to_excel(xl, sheet_name="Resumen", index=False, startrow=2)
        # --- Hoja Diccionario ---
        profile.to_excel(xl, sheet_name="Diccionario", index=False, startrow=1)

        wb = xl.book
        _format_resumen(wb["Resumen"], meta)
        _format_diccionario(wb["Diccionario"], profile)
    return out_path


def _format_resumen(ws, meta: dict) -> None:
    ws["A1"] = "Diccionario de datos — Resumen del dataset"
    ws["A1"].font = _TITLE_FONT
    for r in range(4, 4 + len(meta)):
        ws.cell(r, 1).font = Font(bold=True)
    for col, width in (("A", 28), ("B", 90)):
        ws.column_dimensions[col].width = width
    # encabezado de la mini-tabla
    for c in (1, 2):
        cell = ws.cell(3, c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    ws.cell(3, 1).value, ws.cell(3, 2).value = "Campo", "Valor"
    for row in ws.iter_rows(min_row=4, max_row=3 + len(meta), max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _format_diccionario(ws, profile: pd.DataFrame) -> None:
    ws["A1"] = "Diccionario de datos — Columnas"
    ws["A1"].font = _TITLE_FONT
    widths = {"Columna": 22, "Descripción": 75, "Tipo de dato": 16,
              "% No nulos": 12, "Nº únicos": 11, "Ejemplos": 38, "Observaciones": 28}
    header_row = 2
    for j, name in enumerate(_COLS, start=1):
        cell = ws.cell(header_row, j)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = widths[name]
    ws.freeze_panes = ws.cell(header_row + 1, 1)

    obs_col = _COLS.index("Observaciones") + 1
    pct_col = _COLS.index("% No nulos") + 1
    for i in range(len(profile)):
        r = header_row + 1 + i
        obs = ws.cell(r, obs_col).value or ""
        fill = None
        if "VACÍA" in obs or "Constante" in obs:
            fill = _BAD_FILL
        elif "Mayoría nula" in obs:
            fill = _WARN_FILL
        for j in range(1, len(_COLS) + 1):
            cell = ws.cell(r, j)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(j in (2, 6)))
            if fill is not None:
                cell.fill = fill
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(_COLS))}{header_row + len(profile)}"


def build_meta(name: str, df, source: str, extra: dict | None = None) -> dict:
    """Construye el bloque de metadatos del dataset para la hoja Resumen."""
    n_rows, n_cols = df.shape
    meta = {
        "Dataset": name,
        "Fuente": source,
        "Nº de registros (filas)": f"{n_rows:,}".replace(",", "."),
        "Nº de columnas": n_cols - (1 if "geometry" in getattr(df, "columns", []) else 0),
    }
    if isinstance(df, gpd.GeoDataFrame) and df.geometry.notna().any():
        meta["CRS (sistema de coordenadas)"] = str(df.crs)
        meta["Tipo de geometría"] = ", ".join(df.geom_type.dropna().unique())
        b = [round(float(x), 1) for x in df.total_bounds]
        meta["Extensión (bounds)"] = f"x[{b[0]}, {b[2]}] · y[{b[1]}, {b[3]}]"
    if extra:
        meta.update(extra)
    return meta
